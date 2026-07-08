#!/usr/bin/env python3
"""
Serveur local pour la heatmap animaux sauvages.
Lance avec : python3 server.py
Puis ouvre : http://localhost:8080  (ou http://[IP]:8080 depuis un autre ordi)
"""
import http.server, urllib.request, json, os, threading, time, re, uuid as _uuid, io
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

PORT             = 8080
DEFAULT_SHEET_ID = '1VpN669xp2u34Itapwz5UWuEIi2UmoxTTWfrPQP83rhE'
CONFIG_FILE      = 'config.json'

# ── Configuration persistante (config.json) ────────────────────────────────────
_cfg_lock = threading.Lock()

def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def save_config(cfg):
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)

def get_sheet_id():
    return load_config().get('sheet_id', DEFAULT_SHEET_ID)

def set_sheet_id(new_id):
    with _cfg_lock:
        cfg = load_config()
        cfg['sheet_id'] = new_id
        save_config(cfg)
        # Invalider tous les caches pour forcer un rechargement
        _cache['data']        = None
        _cache['ts']          = 0
        _villes_cache['data'] = None
        _villes_cache['ts']   = 0
        _vetos_cache['data']  = None
        _vetos_cache['ts']    = 0
    print(f'[config] ✓ Sheet ID mis à jour : {new_id}')

# ── Cache données ──────────────────────────────────────────────────────────────
_cache = {'data': None, 'ts': 0}
CACHE_TTL = 300  # secondes (5 min)

def fetch_csv_tab(sheet_id, sheet_name=None):
    """Télécharge un onglet CSV depuis Google Sheets."""
    from urllib.parse import quote
    if sheet_name:
        url = f'https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv&sheet={quote(sheet_name)}'
    else:
        url = f'https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv'
    req = urllib.request.Request(url, headers={'User-Agent': 'heatmap-server/1.0'})
    with urllib.request.urlopen(req, timeout=15) as r:
        return r.read().decode('utf-8')

def get_data():
    now = time.time()
    if _cache['data'] is None or now - _cache['ts'] > CACHE_TTL:
        print('Chargement du Google Sheet...')
        _cache['data'] = fetch_csv_tab(get_sheet_id())
        _cache['ts']   = now
        print(f'  OK — {len(_cache["data"].splitlines())} lignes')
    return _cache['data']

_villes_cache  = {'data': None, 'ts': 0}
_vetos_cache   = {'data': None, 'ts': 0}
_especes_cache = {'data': None, 'ts': 0}

# Statistiques des cliniques (poussées par chaque clinique)
_CLINIC_STATS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'clinic_stats.json')
_clinic_stats_lock = threading.Lock()

def _load_clinic_stats():
    try:
        with open(_CLINIC_STATS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}

def _save_clinic_stats(data):
    try:
        with open(_CLINIC_STATS_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f'Erreur sauvegarde clinic_stats: {e}')

_clinic_stats = _load_clinic_stats()

import secrets as _secrets
import html as _html

# ── Tokens persistants avec expiration (7 jours) ─────────────────────────────
_TOKENS_FILE  = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.auth_tokens.json')
_tokens_lock  = threading.Lock()
TOKEN_TTL     = 7 * 24 * 3600  # 7 jours

def _load_tokens():
    try:
        if os.path.exists(_TOKENS_FILE):
            with open(_TOKENS_FILE, 'r') as f:
                return json.load(f)
    except Exception:
        pass
    return {}

def _save_tokens(tokens):
    with open(_TOKENS_FILE, 'w') as f:
        json.dump(tokens, f)

# Charger les tokens existants au démarrage (survit aux redémarrages)
_auth_tokens = _load_tokens()

def _add_token(token):
    with _tokens_lock:
        _auth_tokens[token] = time.time() + TOKEN_TTL
        _save_tokens(_auth_tokens)

def _check_token(token):
    with _tokens_lock:
        expiry = _auth_tokens.get(token)
        if expiry is None:
            return False
        if time.time() > expiry:
            del _auth_tokens[token]
            _save_tokens(_auth_tokens)
            return False
        return True

def _purge_expired_tokens():
    """Nettoyer les tokens expirés toutes les heures."""
    while True:
        time.sleep(3600)
        now = time.time()
        with _tokens_lock:
            expired = [t for t, exp in list(_auth_tokens.items()) if now > exp]
            for t in expired:
                _auth_tokens.pop(t, None)
            if expired:
                _save_tokens(_auth_tokens)

threading.Thread(target=_purge_expired_tokens, daemon=True).start()

# ── Rate limiting login (5 essais → blocage 15 min) ───────────────────────────
_login_attempts = {}  # ip -> {'count': int, 'blocked_until': float}
_login_lock = threading.Lock()
MAX_LOGIN_ATTEMPTS = 5
BLOCK_DURATION     = 15 * 60  # 15 minutes

_LOCAL_IPS = {'127.0.0.1', '::1', 'localhost'}

def _check_rate_limit(ip):
    """Retourne (allowed: bool, retry_after_seconds: int)."""
    if ip in _LOCAL_IPS:
        return True, 0  # Jamais bloquer le réseau local
    with _login_lock:
        entry = _login_attempts.get(ip, {'count': 0, 'blocked_until': 0})
        now = time.time()
        if now < entry['blocked_until']:
            return False, int(entry['blocked_until'] - now)
        return True, 0

def _record_login_failure(ip):
    with _login_lock:
        entry = _login_attempts.get(ip, {'count': 0, 'blocked_until': 0})
        entry['count'] += 1
        if entry['count'] >= MAX_LOGIN_ATTEMPTS:
            entry['blocked_until'] = time.time() + BLOCK_DURATION
            entry['count'] = 0
            print(f'[auth] IP bloquée 15 min : {ip}')
        _login_attempts[ip] = entry

def _record_login_success(ip):
    with _login_lock:
        _login_attempts.pop(ip, None)

# ── CORS : domaine autorisé ────────────────────────────────────────────────────
def _cors_origin():
    cfg = load_config()
    return cfg.get('allowed_origin', '*')

# ── Sanitisation XSS ──────────────────────────────────────────────────────────
def _sanitize(s, max_len=2000):
    return _html.escape(str(s)[:max_len])

_MESSAGES_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'messages.json')
_messages_lock = threading.Lock()

def _load_messages():
    with _messages_lock:
        if os.path.exists(_MESSAGES_FILE):
            with open(_MESSAGES_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        return []

def _save_messages(msgs):
    with _messages_lock:
        with open(_MESSAGES_FILE, 'w', encoding='utf-8') as f:
            json.dump(msgs, f, ensure_ascii=False, indent=2)

def _check_portal_password(pwd):
    cfg = load_config()
    return pwd == cfg.get('portal_password', 'Hegalaldia2026')

_HISTORY_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'clinic_history.json')
_history_lock = threading.Lock()

def _load_history():
    try:
        with open(_HISTORY_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {'animals_log': [], 'outcomes_log': [], 'logged_animal_ids': [], 'logged_outcome_ids': []}

def _save_history(data):
    try:
        with open(_HISTORY_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f'Erreur sauvegarde history: {e}')

def _log_new_entries(code, incoming):
    """Détecte les nouveaux animaux/sorties et les archive dans l'historique."""
    import datetime as _dt
    now = _dt.datetime.now()
    now_str  = now.strftime('%Y-%m-%d')
    week_str = now.strftime('%Y-W%W')
    with _history_lock:
        history = _load_history()
        logged_animal_ids  = set(history.get('logged_animal_ids', []))
        logged_outcome_ids = set(history.get('logged_outcome_ids', []))
        changed = False
        for a in incoming.get('animalsList', []):
            aid = str(a.get('id', ''))
            if aid and aid not in logged_animal_ids:
                history['animals_log'].append({
                    'id': aid, 'date': now_str, 'week': week_str,
                    'species': a.get('species', ''), 'commune': a.get('commune', ''),
                    'clinic_code': code
                })
                logged_animal_ids.add(aid)
                changed = True
        for o in incoming.get('outcomesList', []):
            oid = str(o.get('id', ''))
            if oid and oid not in logged_outcome_ids:
                history['outcomes_log'].append({
                    'id': oid, 'date': now_str, 'week': week_str,
                    'species': o.get('species', ''), 'type': o.get('type', 'deces'),
                    'commune': o.get('commune', ''), 'clinic_code': code
                })
                logged_outcome_ids.add(oid)
                changed = True
        if changed:
            history['logged_animal_ids']  = list(logged_animal_ids)
            history['logged_outcome_ids'] = list(logged_outcome_ids)
            _save_history(history)

def _weekly_clear_sorties():
    """Vide les sorties (outcomesList + deces) tous les dimanches à 23h."""
    import datetime as _dt
    while True:
        now = _dt.datetime.now()
        # Calculer le prochain dimanche 23:00
        days_until_sunday = (6 - now.weekday()) % 7  # 6 = dimanche
        next_sunday = now.replace(hour=23, minute=0, second=0, microsecond=0) + _dt.timedelta(days=days_until_sunday)
        if next_sunday <= now:
            next_sunday += _dt.timedelta(weeks=1)
        wait_seconds = (next_sunday - now).total_seconds()
        print(f'[cron] Prochain effacement des sorties : {next_sunday.strftime("%A %d/%m à %H:%M")} (dans {int(wait_seconds//3600)}h)')
        time.sleep(wait_seconds)
        # Snapshot hebdomadaire avant effacement
        import datetime as _dt2
        week_label = _dt2.datetime.now().strftime('%Y-W%W')
        with _history_lock:
            h = _load_history()
            # Marquer tous les outcomes de cette semaine comme "archivés" (ils le sont déjà dans outcomes_log)
            _save_history(h)
        # Effacer les sorties de toutes les cliniques
        with _clinic_stats_lock:
            for code in _clinic_stats:
                _clinic_stats[code]['outcomesList'] = []
                _clinic_stats[code]['deces'] = 0
            _save_clinic_stats(_clinic_stats)
        print(f'[cron] ✓ Sorties effacées ({_dt.datetime.now().strftime("%d/%m/%Y %H:%M")})')

_cron_thread = threading.Thread(target=_weekly_clear_sorties, daemon=True)
_cron_thread.start()

# ── Sauvegarde automatique journalière ────────────────────────────────────────
_BACKUP_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'backups')

def _daily_backup():
    """Sauvegarde journalière de tous les JSON à 3h du matin."""
    import datetime as _dt
    os.makedirs(_BACKUP_DIR, exist_ok=True)
    while True:
        now = _dt.datetime.now()
        next_3am = now.replace(hour=3, minute=0, second=0, microsecond=0)
        if next_3am <= now:
            next_3am += _dt.timedelta(days=1)
        wait_seconds = (next_3am - now).total_seconds()
        print(f'[backup] Prochain backup : {next_3am.strftime("%d/%m à %H:%M")} (dans {int(wait_seconds//3600)}h)')
        time.sleep(wait_seconds)
        label = _dt.datetime.now().strftime('%Y-%m-%d')
        files = {
            'clinic_stats': _CLINIC_STATS_FILE,
            'clinic_history': _HISTORY_FILE,
            'messages': _MESSAGES_FILE,
            'benevoles': BENEVOLES_JSON if 'BENEVOLES_JSON' in dir() else None,
        }
        count = 0
        for name, path in files.items():
            if path and os.path.exists(path):
                import shutil
                dest = os.path.join(_BACKUP_DIR, f'{label}_{name}.json')
                shutil.copy2(path, dest)
                count += 1
        # Garder seulement les 30 derniers jours de backups
        all_backups = sorted(
            [f for f in os.listdir(_BACKUP_DIR) if f.endswith('.json')],
            reverse=True
        )
        seen_days = set()
        for fname in all_backups:
            day = fname[:10]
            seen_days.add(day)
            if len(seen_days) > 30:
                os.remove(os.path.join(_BACKUP_DIR, fname))
        print(f'[backup] ✓ {count} fichiers sauvegardés ({label})')

_backup_thread = threading.Thread(target=_daily_backup, daemon=True)
_backup_thread.start()

def get_villes():
    now = time.time()
    if _villes_cache['data'] is None or now - _villes_cache['ts'] > CACHE_TTL:
        print('Chargement onglet Ville...')
        csv = fetch_csv_tab(get_sheet_id(), 'Ville')
        # Extraire uniquement les noms de communes (colonne "Ville de decouverte")
        import csv as csvmod, io
        reader = csvmod.DictReader(io.StringIO(csv))
        villes = sorted(set(
            r['Ville de decouverte'].strip()
            for r in reader
            if r.get('Ville de decouverte', '').strip()
        ))
        _villes_cache['data'] = villes
        _villes_cache['ts']   = now
        print(f'  OK — {len(villes)} communes')
    return _villes_cache['data']

import unicodedata
def strip_accents(s):
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')

def get_especes():
    now = time.time()
    if _especes_cache['data'] is None or now - _especes_cache['ts'] > CACHE_TTL:
        print('Chargement onglet Divers (espèces)...')
        import csv as csvmod, io
        csv_text = fetch_csv_tab(get_sheet_id(), 'Divers')
        reader = csvmod.reader(io.StringIO(csv_text))
        especes = sorted(set(
            row[0].strip()
            for row in reader
            if row and row[0].strip()
        ))
        _especes_cache['data'] = especes
        _especes_cache['ts']   = now
        print(f'  OK — {len(especes)} espèces')
    return _especes_cache['data']

def get_benevoles_from_xls():
    """Lit le fichier listingbenevole.xls (format HTML) et retourne la liste des bénévoles."""
    from html.parser import HTMLParser
    class TableParser(HTMLParser):
        def __init__(self):
            super().__init__()
            self.rows=[]; self.current_row=[]; self.current_cell=''; self.in_cell=False
        def handle_starttag(self,tag,attrs):
            if tag in ('td','th'): self.in_cell=True; self.current_cell=''
            elif tag=='tr': self.current_row=[]
        def handle_endtag(self,tag):
            if tag in ('td','th'): self.current_row.append(self.current_cell.strip()); self.in_cell=False
            elif tag=='tr':
                if any(c.strip() for c in self.current_row): self.rows.append(self.current_row)
        def handle_data(self,data):
            if self.in_cell: self.current_cell+=data
    xls_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'listingbenevole.xls')
    with open(xls_path, encoding='utf-8') as f:
        content = f.read()
    p = TableParser(); p.feed(content)
    if not p.rows: return []
    headers = p.rows[0]
    result = []
    for row in p.rows[1:]:
        # Compléter si la ligne est plus courte que les headers
        while len(row) < len(headers): row.append('')
        d = {headers[i]: row[i] for i in range(len(headers))}
        # Déterminer le(s) rôle(s)
        roles = []
        if d.get('Bénévole sur centre','').strip(): roles.append('centre')
        if d.get('Bénévole rapatrieur','').strip(): roles.append('rapatrieur')
        if d.get('Autre bénévole','').strip(): roles.append('autre')
        if d.get('À former','').strip(): roles.append('former')
        prenom = d.get('Prénom','').strip()
        nom    = d.get('Nom','').strip()
        if not prenom and not nom: continue
        result.append({
            'prenom':      prenom,
            'nom':         nom,
            'roles':       roles,
            'telephone':   d.get('Téléphone(s)','').strip(),
            'email':       d.get('Email','').strip(),
            'commune':     d.get("Commune d'habitation",'').strip(),
            'rayon':       d.get('Rayon intervention (en kms)','').strip(),
            'naissance':   d.get('Date de naissance','').strip(),
            'immat':       d.get('Immatriculation du véhicule','').strip(),
            'puissance':   d.get('Puissance du véhicule','').strip(),
            'moteur':      d.get('Moteur du véhicule','').strip(),
            'competences': d.get('Compétences','').strip(),
            'commentaire': d.get('Commentaire','').strip(),
            'info_temp':   d.get('Information temporaire','').strip(),
            'derniers_rapat':  d.get('Derniers rapatriements','').strip(),
            'date_creation':   d.get('Date de création','').strip(),
            'modif':           d.get('Dernière modification','').strip(),
            'devenir':         d.get('Devenir','').strip(),
            'veterinaire':     d.get('Vétérinaire','').strip(),
        })
    return result

BENEVOLES_JSON = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'benevoles.json')
_benv_lock = threading.RLock()  # RLock = réentrant, évite deadlock si get_benevoles() appelé dans un with _benv_lock

def load_benevoles_json():
    """Returns list if JSON exists, None if not."""
    if os.path.exists(BENEVOLES_JSON):
        try:
            with open(BENEVOLES_JSON, encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return []
    return None

def save_benevoles_json(data):
    with open(BENEVOLES_JSON, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

EXTRA_FIELDS = [
    'telephone2', 'adresse', 'urg_nom', 'urg_prenom', 'urg_tel',
    'forme', 'forme_soin', 'forme_rapat',
    'adhesion', 'admin', 'ca', 'veteran', 'anc_sc', 'anc_salarie', 'anc_ca',
    'premier_jour', 'disponibilites',
    'comp_menage', 'comp_oisillons', 'comp_marins', 'comp_vautour', 'comp_exterieurs', 'comp_heris',
    'caisses', 'caisses_nb', 'caisses_taille', 'gants', 'garde_domicile', 'recu_fiscal',
    'zone_pb', 'zone_bearn', 'zone_landes', 'zone_hp',
    'comp_brico', 'comp_info', 'comp_commu', 'comp_compta', 'comp_photo', 'comp_couture',
    'act_caddie', 'act_papier_cadeau', 'act_espaces_verts',
    'autres_competences',
    'devenir', 'veterinaire',
]

def import_xls_to_json():
    """Read XLS and create benevoles.json. Called only when JSON doesn't exist."""
    volunteers = get_benevoles_from_xls()
    for b in volunteers:
        b['_id'] = str(_uuid.uuid4())
        for k in EXTRA_FIELDS:
            b.setdefault(k, '')
    save_benevoles_json(volunteers)
    print(f'[benevoles] ✓ Import XLS → JSON : {len(volunteers)} bénévoles')
    return volunteers

def get_benevoles():
    with _benv_lock:
        data = load_benevoles_json()
        if data is None:
            data = import_xls_to_json()
        return data

def add_benevole(data):
    from datetime import datetime
    now = datetime.now().strftime('%d/%m/%y à %H:%M')
    roles = [r for r in ['centre', 'rapatrieur', 'autre', 'former'] if data.get(r)]
    new_b = {
        '_id': str(_uuid.uuid4()),
        'prenom': data.get('prenom', ''),
        'nom':    data.get('nom', ''),
        'roles':  roles,
        'date_creation': now,
        'modif': now,
        'derniers_rapat': '',
    }
    all_fields = [
        'telephone', 'telephone2', 'email', 'adresse', 'commune', 'rayon', 'naissance',
        'urg_nom', 'urg_prenom', 'urg_tel',
        'forme', 'forme_soin', 'forme_rapat',
        'adhesion', 'admin', 'ca', 'veteran', 'anc_sc', 'anc_salarie', 'anc_ca',
        'premier_jour', 'disponibilites',
        'comp_menage', 'comp_oisillons', 'comp_marins', 'comp_vautour', 'comp_exterieurs', 'comp_heris',
        'immat', 'puissance', 'moteur',  # legacy
        'caisses', 'caisses_nb', 'caisses_taille', 'gants', 'garde_domicile', 'recu_fiscal',
        'zone_pb', 'zone_bearn', 'zone_landes', 'zone_hp',
        'comp_brico', 'comp_info', 'comp_commu', 'comp_compta', 'comp_photo', 'comp_couture',
        'act_caddie', 'act_papier_cadeau', 'act_espaces_verts',
        'autres_competences', 'commentaire', 'info_temp', 'competences',
    ]
    for k in all_fields:
        new_b[k] = data.get(k, '')
    # Nouveau champ vehicules (tableau)
    new_b['vehicules'] = data.get('vehicules', [])
    with _benv_lock:
        volunteers = get_benevoles()
        nom_new    = new_b['nom'].strip().lower()
        prenom_new = new_b['prenom'].strip().lower()
        for b in volunteers:
            if b.get('nom','').strip().lower() == nom_new and b.get('prenom','').strip().lower() == prenom_new:
                raise ValueError(f'Un bénévole nommé {new_b["prenom"]} {new_b["nom"]} existe déjà.')
        volunteers.append(new_b)
        save_benevoles_json(volunteers)
    print(f'[benevoles] ✓ Ajouté : {new_b["prenom"]} {new_b["nom"]}')

def update_benevole(orig_nom, orig_prenom, data):
    from datetime import datetime
    now = datetime.now().strftime('%d/%m/%y à %H:%M')
    roles = [r for r in ['centre', 'rapatrieur', 'autre', 'former'] if data.get(r)]
    with _benv_lock:
        volunteers = get_benevoles()
        for i, b in enumerate(volunteers):
            if b.get('nom','').strip() == orig_nom.strip() and b.get('prenom','').strip() == orig_prenom.strip():
                preserved = {
                    '_id':            b.get('_id', str(_uuid.uuid4())),
                    'derniers_rapat': b.get('derniers_rapat', ''),
                    'date_creation':  b.get('date_creation', ''),
                }
                updated = {
                    **preserved,
                    'prenom': data.get('prenom', orig_prenom),
                    'nom':    data.get('nom', orig_nom),
                    'roles':  roles,
                    'modif':  now,
                }
                all_fields = [
                    'telephone', 'telephone2', 'email', 'adresse', 'commune', 'rayon', 'naissance',
                    'urg_nom', 'urg_prenom', 'urg_tel',
                    'forme', 'forme_soin', 'forme_rapat',
                    'adhesion', 'admin', 'ca', 'veteran', 'anc_sc', 'anc_salarie', 'anc_ca',
                    'premier_jour', 'disponibilites',
                    'comp_menage', 'comp_oisillons', 'comp_marins', 'comp_vautour', 'comp_exterieurs', 'comp_heris',
                    'immat', 'puissance', 'moteur',
                    'caisses', 'caisses_nb', 'caisses_taille', 'gants', 'garde_domicile', 'recu_fiscal',
                    'zone_pb', 'zone_bearn', 'zone_landes', 'zone_hp',
                    'act_caddie', 'act_papier_cadeau', 'act_espaces_verts',
                    'comp_brico', 'comp_info', 'comp_commu', 'comp_compta', 'comp_photo', 'comp_couture',
                    'autres_competences', 'commentaire', 'info_temp', 'competences',
                ]
                for k in all_fields:
                    updated[k] = data.get(k, b.get(k, ''))
                updated['vehicules'] = data.get('vehicules', b.get('vehicules', []))
                # info_temp is legacy; commentaire replaces it — always clear on update
                updated['info_temp'] = ''
                volunteers[i] = updated
                save_benevoles_json(volunteers)
                print(f'[benevoles] ✓ Modifié : {orig_prenom} {orig_nom}')
                return
    raise ValueError(f'Bénévole non trouvé : {orig_prenom} {orig_nom}')

_CLINIC_CODES_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'clinic_codes.json')
_clinic_codes_lock = threading.Lock()

def _load_clinic_codes():
    try:
        with open(_CLINIC_CODES_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}

def _save_clinic_codes(data):
    try:
        with open(_CLINIC_CODES_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f'Erreur sauvegarde clinic_codes: {e}')

def _generate_clinic_code(nom, existing_codes):
    """Génère un code unique et déterministe pour une clinique."""
    import unicodedata
    # Préfixe : 3 premières lettres de la ville (avant la parenthèse), sans accents
    city = nom.split('(')[0].strip()
    city_clean = ''.join(
        c for c in unicodedata.normalize('NFD', city)
        if unicodedata.category(c) != 'Mn' and c.isalpha()
    ).upper()
    prefix = city_clean[:3] if len(city_clean) >= 3 else city_clean.ljust(3, 'X')
    # Numéro : hash déterministe du nom complet → 100-999
    h = 0
    for ch in nom:
        h = (h * 31 + ord(ch)) & 0xFFFF
    base_num = 100 + (h % 900)
    # Éviter les collisions avec les codes déjà attribués
    used_nums = {int(c.split('-')[1]) for c in existing_codes.values()
                 if c.startswith(prefix + '-') and c.split('-')[1].isdigit()}
    num = base_num
    while num in used_nums:
        num = 100 + (num + 1) % 900
    return f'{prefix}-{num:03d}'

def get_vetos():
    now = time.time()
    if _vetos_cache['data'] is None or now - _vetos_cache['ts'] > CACHE_TTL:
        print('Chargement onglet Vétérinaire...')
        csv = fetch_csv_tab(get_sheet_id(), 'Vétérinaires')
        import csv as csvmod, io
        reader = csvmod.reader(io.StringIO(csv))
        rows   = list(reader)
        vetos  = []
        for row in rows[1:]:   # ignorer la ligne d'en-tête
            if not row or not row[0].strip():
                continue
            nom    = row[0].strip()
            statut = row[1].strip() if len(row) > 1 else ''
            statut_norm = statut.lower()
            if 'partenaire' in statut_norm:
                statut = 'Partenaire'
            elif 'eviter' in statut_norm or 'éviter' in statut_norm:
                statut = 'Eviter'
            else:
                statut = ''
            vetos.append({'nom': nom, 'statut': statut})
        # Assigner/compléter les codes des cliniques partenaires
        with _clinic_codes_lock:
            codes = _load_clinic_codes()
            changed = False
            for v in vetos:
                if v['statut'] == 'Partenaire' and v['nom'] not in codes:
                    codes[v['nom']] = _generate_clinic_code(v['nom'], codes)
                    print(f'  [codes] Nouveau code : {v["nom"]} → {codes[v["nom"]]}')
                    changed = True
                if v['statut'] == 'Partenaire':
                    v['code'] = codes[v['nom']]
            if changed:
                _save_clinic_codes(codes)
        _vetos_cache['data'] = vetos
        _vetos_cache['ts']   = now
        print(f'  OK — {len(vetos)} vétérinaires ({sum(1 for v in vetos if v["statut"]=="Partenaire")} partenaires)')
    return _vetos_cache['data']

# ── Géocache HTML ──────────────────────────────────────────────────────────────
HTML_FILE = 'heatmap.html'
GEO_LOCK  = threading.Lock()

def add_to_geo_preloaded(ville, lat, lon):
    """Ajoute une ville dans GEO_PRELOADED directement dans le fichier HTML."""
    with GEO_LOCK:
        with open(HTML_FILE, 'r', encoding='utf-8') as f:
            html = f.read()

        m = re.search(r'const GEO_PRELOADED = (\{.*?\});', html, re.DOTALL)
        if not m:
            return False, 'GEO_PRELOADED introuvable dans le HTML'

        geo = json.loads(m.group(1))
        if ville in geo:
            return True, 'déjà présent'

        geo[ville] = {'lat': lat, 'lon': lon}
        new_json  = json.dumps(geo, ensure_ascii=False, separators=(',', ':'))
        new_html  = html[:m.start(1)] + new_json + html[m.end(1):]

        with open(HTML_FILE, 'w', encoding='utf-8') as f:
            f.write(new_html)

        print(f'[geocache] ✓ Ajouté : {ville} ({lat}, {lon})')
        return True, 'ajouté'


# ── Export XLS ────────────────────────────────────────────────────────────────
def export_benevoles_xls(year=None):
    data = get_benevoles()
    # Si year spécifié : filtre les rapatriements pour n'inclure que cette année
    if year:
        import copy
        data = copy.deepcopy(data)
        for b in data:
            raw = b.get('derniers_rapat', '')
            rows = []
            if isinstance(raw, list):
                rows = raw
            elif raw:
                for ligne in str(raw).split('\n'):
                    ligne = ligne.strip()
                    if not ligne: continue
                    if ' : ' in ligne:
                        d, t = ligne.split(' : ', 1)
                    else:
                        d, t = '', ligne
                    rows.append({'date': d.strip(), 'trajet': t.strip()})
            kept = [r for r in rows if rapat_year(r) == year]
            if isinstance(raw, list):
                b['derniers_rapat'] = kept
            else:
                b['derniers_rapat'] = '\n'.join(
                    (f"{r['date']} : {r.get('trajet','')}" if r.get('date') else r.get('trajet','')).strip()
                    for r in kept
                )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Bénévoles'

    # ── Couleurs ──
    C_GREEN  = 'FF0C3D22'
    C_LGREE  = 'FFE8F5E9'
    C_ORANG  = 'FFEC6724'
    C_LORAN  = 'FFFFF3EC'
    C_GRAY   = 'FF5A6672'
    C_LGRAY  = 'FFF2F4F5'
    C_WHITE  = 'FFFFFFFF'
    C_HEAD   = 'FF263238'

    def hfont(color='FFFFFFFF', bold=True, sz=10):
        return Font(name='Calibri', bold=bold, color=color, size=sz)
    def cell_fill(color):
        return PatternFill('solid', fgColor=color)
    def thin_border(color='FFCCCCCC'):
        s = Side(style='thin', color=color)
        return Border(left=s, right=s, top=s, bottom=s)
    def center():
        return Alignment(horizontal='center', vertical='center', wrap_text=True)
    def left():
        return Alignment(horizontal='left', vertical='center', wrap_text=True)

    # ── Colonnes groupées ──
    COLS = [
        # (header, champ_json, largeur, groupe)
        # — Identité —
        ('Nom',           'nom',         18, 'id'),
        ('Prénom',        'prenom',       16, 'id'),
        ('Naissance',     'naissance',    14, 'id'),
        ('Commune',       'commune',      18, 'id'),
        ('Adresse',       'adresse',      28, 'id'),
        # — Contact —
        ('Téléphone',     'telephone',    16, 'contact'),
        ('Téléphone 2',   'telephone2',   16, 'contact'),
        ('Email',         'email',        28, 'contact'),
        # — Contact urgence —
        ('Urg. Nom',      'urg_nom',      16, 'urg'),
        ('Urg. Prénom',   'urg_prenom',   16, 'urg'),
        ('Urg. Tél.',     'urg_tel',      16, 'urg'),
        # — Rôles —
        ('Sur centre',    'centre',       12, 'role'),
        ('Rapatrieur',    'rapatrieur',   12, 'role'),
        ('Autre',         'autre',        10, 'role'),
        # — Formation —
        ('Formé',         'forme',        10, 'form'),
        ('Formé soin',    'forme_soin',   12, 'form'),
        ('Formé rapat.',  'forme_rapat',  12, 'form'),
        # — Sur centre —
        ('Adhésion',      'adhesion',     10, 'centre'),
        ('1er jour',      'premier_jour', 16, 'centre'),
        ('Disponibilités','disponibilites',28,'centre'),
        ('Comp. Ménage',  'comp_menage',  14, 'centre'),
        ('Comp. Oisillons','comp_oisillons',14,'centre'),
        ('Comp. Marins',  'comp_marins',  14, 'centre'),
        ('Comp. Vautour', 'comp_vautour', 14, 'both'),
        ('Comp. Extérieurs','comp_exterieurs',14,'centre'),
        ('Comp. Hérissons','comp_heris',  14, 'centre'),
        # — Rapatrieur —
        ('Immatriculation','immat',       16, 'rapat'),
        ('Puissance (CV)', 'puissance',   14, 'rapat'),
        ('Moteur',        'moteur',       12, 'rapat'),
        ('Caisses transp.','caisses',     12, 'rapat'),
        ('Nb caisses',    'caisses_nb',   10, 'rapat'),
        ('Taille caisses','caisses_taille',14,'rapat'),
        ('Gants',         'gants',        10, 'rapat'),
        ('Garde domicile','garde_domicile',14,'rapat'),
        ('Reçu fiscal',   'recu_fiscal',  12, 'rapat'),
        ('Zone PB',       'zone_pb',      10, 'rapat'),
        ('Zone Béarn',    'zone_bearn',   12, 'rapat'),
        ('Zone Landes',   'zone_landes',  12, 'rapat'),
        ('Zone HP',       'zone_hp',      10, 'rapat'),
        ('Rayon (km)',    'rayon',        12, 'rapat'),
        # — Autre —
        ('Comp. Bricolage','comp_brico',  14, 'autre'),
        ('Comp. Informatique','comp_info',14, 'autre'),
        ('Comp. Communication','comp_commu',14,'autre'),
        ('Comp. Comptabilité','comp_compta',14,'autre'),
        ('Comp. Photo',   'comp_photo',   12, 'autre'),
        ('Comp. Couture', 'comp_couture', 12, 'autre'),
        ('Autres comp.',  'autres_competences',28,'autre'),
        # — Historique —
        ('Rapatriements', 'derniers_rapat',30,'hist'),
        ('Commentaire',   'commentaire',  36, 'hist'),
        ('Bénévole depuis','date_creation',18,'hist'),
        ('Dernière modif.','modif',        18,'hist'),
    ]

    GROUP_LABELS = {
        'id':      ('Identité',           C_HEAD),
        'contact': ('Contact',            C_GREEN),
        'urg':     ('Urgence',            C_GRAY),
        'role':    ('Rôles',              C_HEAD),
        'form':    ('Formation',          C_HEAD),
        'centre':  ('Sur centre',         '  FF019939'),
        'rapat':   ('Rapatrieur',         C_ORANG),
        'autre':   ('Autre / Compétences',C_GRAY),
        'hist':    ('Historique',         C_HEAD),
    }

    # ── Ligne 1 : groupes (merged cells) ──
    cur_col = 1
    group_ranges = {}
    for _, _, _, grp in COLS:
        if grp not in group_ranges:
            group_ranges[grp] = [cur_col, cur_col]
        else:
            group_ranges[grp][1] = cur_col
        cur_col += 1

    for grp, (start, end) in group_ranges.items():
        label, color = GROUP_LABELS[grp]
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        c = ws.cell(row=1, column=start, value=label)
        c.font      = hfont(sz=9)
        c.fill      = cell_fill(color.replace('  ', ''))
        c.alignment = center()

    # ── Ligne 2 : en-têtes colonnes ──
    for col_idx, (header, _, width, grp) in enumerate(COLS, start=1):
        _, color = GROUP_LABELS[grp]
        # Couleur légèrement plus claire
        light = {'id': C_LGRAY, 'contact': C_LGREE, 'urg': C_LGRAY,
                  'role': C_LGRAY, 'form': C_LGRAY, 'centre': 'FFE8F5E9',
                  'rapat': C_LORAN, 'autre': C_LGRAY, 'hist': C_LGRAY}.get(grp, C_LGRAY)
        c = ws.cell(row=2, column=col_idx, value=header)
        c.font      = Font(name='Calibri', bold=True, size=9, color='FF263238')
        c.fill      = cell_fill(light)
        c.alignment = center()
        c.border    = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 30
    ws.freeze_panes = 'A3'

    # ── Données ──
    def parse_rapat(val):
        """
        Accepte soit une liste d'objets {date, animal, lieu, ...}
        soit un texte brut multi-lignes "06/01/26 : Pau > Ustaritz".
        Retourne toujours une liste de dicts {date, trajet, animal, lieu}.
        """
        if not val:
            return []
        if isinstance(val, list):
            return val
        # Texte brut : une entrée par ligne
        lignes = [l.strip() for l in str(val).split('\n') if l.strip()]
        result = []
        for ligne in lignes:
            # Format attendu : "DATE : TRAJET" ou juste "TRAJET"
            if ' : ' in ligne:
                date_part, trajet = ligne.split(' : ', 1)
            else:
                date_part, trajet = '', ligne
            result.append({'date': date_part.strip(), 'trajet': trajet.strip()})
        return result

    def fmt_rapat(val):
        rows = parse_rapat(val)
        if not rows:
            return ''
        lines = []
        for r in rows:
            line = r.get('date', '')
            trajet_raw = r.get('trajet', '') or r.get('lieu', '') or r.get('commune', '')
            trajet = ' > '.join(trajet_raw) if isinstance(trajet_raw, list) else trajet_raw
            dist   = r.get('distance_km', '')
            animal = r.get('animal', r.get('espece', ''))
            if trajet: line += (' : ' if line else '') + trajet
            if dist:   line += f' ({dist} km)'
            if animal: line += f' [{animal}]'
            lines.append(line)
        return '\n'.join(lines)

    def yn(val):
        """Convertit 'on'/'' en Oui/Non."""
        if val == 'on': return 'Oui'
        return ''

    for row_idx, b in enumerate(data, start=3):
        fill_col = C_LGREE if b.get('centre') == 'on' else (C_LORAN if b.get('rapatrieur') == 'on' else C_WHITE)
        is_alt   = row_idx % 2 == 0

        # Résoudre les véhicules (nouveau format ou legacy)
        vehs = b.get('vehicules') or []
        if not vehs and (b.get('immat') or b.get('puissance') or b.get('moteur')):
            vehs = [{'immat': b.get('immat',''), 'puissance': b.get('puissance',''), 'moteur': b.get('moteur','')}]

        for col_idx, (_, field, _, grp) in enumerate(COLS, start=1):
            if field == 'immat':
                val = ' | '.join(v.get('immat','') for v in vehs if v.get('immat'))
            elif field == 'puissance':
                val = ' | '.join(v.get('puissance','') for v in vehs if v.get('puissance'))
            elif field == 'moteur':
                val = ' | '.join(v.get('moteur','') for v in vehs if v.get('moteur'))
            else:
                val = b.get(field, '')

            # Mise en forme selon le champ
            if field in ('centre', 'rapatrieur', 'autre',
                         'forme', 'forme_soin', 'forme_rapat',
                         'adhesion', 'caisses', 'gants', 'garde_domicile', 'recu_fiscal',
                         'zone_pb', 'zone_bearn', 'zone_landes', 'zone_hp',
                         'comp_menage', 'comp_oisillons', 'comp_marins', 'comp_vautour', 'comp_exterieurs', 'comp_heris',
                         'comp_brico', 'comp_info', 'comp_commu', 'comp_compta', 'comp_photo', 'comp_couture'):
                val = yn(val)
            elif field == 'derniers_rapat':
                val = fmt_rapat(val) if isinstance(val, list) else (val or '')
            elif field in ('date_creation', 'modif'):
                val = str(val or '').split(' ')[0] if val else ''
            else:
                val = str(val or '')

            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font      = Font(name='Calibri', size=9)
            c.alignment = left() if col_idx > 3 else center()
            c.border    = thin_border('FFEEEEEE')
            if is_alt:
                c.fill = cell_fill('FFF9F9F9')

        ws.row_dimensions[row_idx].height = 15

    # ── Feuille résumé rapatriements ──
    ws2 = wb.create_sheet('Rapatriements')
    rap_headers = ['Nom', 'Prénom', 'Date', 'H. départ', 'H. arrivée', 'Distance (km)', 'Trajet / Détail', 'Animal / Espèce', 'Notes']
    rap_widths  = [18, 16, 14, 12, 12, 14, 40, 22, 30]
    for ci, (h, w) in enumerate(zip(rap_headers, rap_widths), 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font  = Font(name='Calibri', bold=True, size=9, color='FFFFFFFF')
        c.fill  = cell_fill(C_ORANG)
        c.alignment = center()
        c.border = thin_border()
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.row_dimensions[1].height = 22
    ws2.freeze_panes = 'A2'

    rap_row = 2
    for b in data:
        rapat_list = parse_rapat(b.get('derniers_rapat', ''))
        for r in rapat_list:
            # Trajet : liste ou texte brut
            trajet_raw = r.get('trajet', '')
            if isinstance(trajet_raw, list):
                trajet = ' > '.join(trajet_raw)
            else:
                trajet = trajet_raw or r.get('lieu', '') or r.get('commune', '')
            animal   = r.get('animal', r.get('espece', ''))
            notes    = r.get('notes', r.get('commentaire', ''))
            dist     = r.get('distance_km', '')
            h_dep    = r.get('heure_depart', '')
            h_arr    = r.get('heure_arrivee', '')
            # Distance en nombre si possible
            try:
                dist_val = int(dist) if dist else ''
            except (ValueError, TypeError):
                dist_val = dist
            ws2.cell(rap_row, 1, b.get('nom', ''))
            ws2.cell(rap_row, 2, b.get('prenom', ''))
            ws2.cell(rap_row, 3, r.get('date', ''))
            ws2.cell(rap_row, 4, h_dep)
            ws2.cell(rap_row, 5, h_arr)
            ws2.cell(rap_row, 6, dist_val)
            ws2.cell(rap_row, 7, trajet)
            ws2.cell(rap_row, 8, animal)
            ws2.cell(rap_row, 9, notes)
            for ci in range(1, 10):
                c = ws2.cell(rap_row, ci)
                c.font      = Font(name='Calibri', size=9)
                c.alignment = left()
                c.border    = thin_border('FFEEEEEE')
                if rap_row % 2 == 0:
                    c.fill = cell_fill('FFFFF8F4')
            # Colonne distance centrée
            ws2.cell(rap_row, 6).alignment = center()
            ws2.row_dimensions[rap_row].height = 14
            rap_row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def rapat_year(r):
    """Extrait l'année d'un rapatriement (dict ou ligne texte parsée)."""
    date_str = r.get('date', '') if isinstance(r, dict) else ''
    if not date_str:
        return None
    parts = date_str.replace('-', '/').split('/')
    if len(parts) < 3:
        return None
    y = parts[2].strip()
    if len(y) == 2:
        y = '20' + y
    try:
        return int(y)
    except ValueError:
        return None


def clear_rapatriements_year(year):
    """Supprime du JSON tous les rapatriements dont l'année == year."""
    with _benv_lock:
        data = load_benevoles_json() or []
        cleared = 0
        for b in data:
            raw = b.get('derniers_rapat', '')
            rows = []
            if isinstance(raw, list):
                rows = raw
            elif raw:
                # Texte brut → parse
                for ligne in str(raw).split('\n'):
                    ligne = ligne.strip()
                    if not ligne:
                        continue
                    if ' : ' in ligne:
                        d, t = ligne.split(' : ', 1)
                    else:
                        d, t = '', ligne
                    rows.append({'date': d.strip(), 'trajet': t.strip()})

            before = len(rows)
            kept   = [r for r in rows if rapat_year(r) != year]
            cleared += before - len(kept)

            # Reconvertit dans le même format qu'avant (texte brut si c'était texte)
            if isinstance(raw, list):
                b['derniers_rapat'] = kept
            else:
                b['derniers_rapat'] = '\n'.join(
                    (f"{r['date']} : {r.get('trajet','')}" if r.get('date') else r.get('trajet', '')).strip()
                    for r in kept
                )
        save_benevoles_json(data)
    return {'cleared': cleared, 'year': year}


# ── Serveur HTTP ───────────────────────────────────────────────────────────────
class Handler(http.server.SimpleHTTPRequestHandler):

    def do_GET(self):
        if self.path == '/api/data':
            try:
                csv  = get_data()
                body = csv.encode('utf-8')
                self.send_response(200)
                self.send_header('Content-Type', 'text/csv; charset=utf-8')
                self.send_header('Content-Length', len(body))
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(body)
            except Exception as e:
                self.send_response(502)
                self.send_header('Content-Type', 'text/plain')
                self.end_headers()
                self.wfile.write(str(e).encode())

        elif self.path.startswith('/api/villes'):
            try:
                villes = get_villes()
                # Filtre optionnel ?q=...
                q = ''
                if '?q=' in self.path:
                    q = self.path.split('?q=')[1].split('&')[0].lower()
                if q:
                    from urllib.parse import unquote
                    q = unquote(q).lower()
                    # Priorité : commence par q d'abord, puis contient q
                    qn = strip_accents(q)
                    starts = [v for v in villes if strip_accents(v.lower()).startswith(qn)][:8]
                    contains = [v for v in villes if qn in strip_accents(v.lower()) and v not in starts][:max(0, 8-len(starts))]
                    result = starts + contains
                else:
                    result = villes
                body = json.dumps(result, ensure_ascii=False).encode('utf-8')
                self.send_response(200)
                self.send_header('Content-Type', 'application/json; charset=utf-8')
                self.send_header('Content-Length', len(body))
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(body)
            except Exception as e:
                self.send_response(502)
                self.send_header('Content-Type', 'text/plain')
                self.end_headers()
                self.wfile.write(str(e).encode())

        elif self.path.startswith('/api/especes'):
            try:
                especes = get_especes()
                q = ''
                if '?q=' in self.path:
                    from urllib.parse import unquote
                    q = unquote(self.path.split('?q=')[1].split('&')[0]).lower()
                if q:
                    qn = strip_accents(q)
                    starts   = [e for e in especes if strip_accents(e.lower()).startswith(qn)][:10]
                    contains = [e for e in especes if qn in strip_accents(e.lower()) and e not in starts][:max(0, 10-len(starts))]
                    result = starts + contains
                else:
                    result = especes
                body = json.dumps(result, ensure_ascii=False).encode('utf-8')
                self.send_response(200)
                self.send_header('Content-Type', 'application/json; charset=utf-8')
                self.send_header('Content-Length', len(body))
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(body)
            except Exception as e:
                self.send_response(502)
                self.send_header('Content-Type', 'text/plain')
                self.end_headers()
                self.wfile.write(str(e).encode())

        elif self.path == '/api/rapatriements/en-route':
            # Retourne {destination: count} pour les rapatriements du jour
            try:
                import datetime as _dt
                today_iso = _dt.date.today().isoformat()
                counts = {}
                volunteers = load_benevoles_json() or []
                for b in volunteers:
                    raw = b.get('derniers_rapat', '')
                    rows = raw if isinstance(raw, list) else []
                    for r in rows:
                        if not isinstance(r, dict):
                            continue
                        dates = r.get('dates', [])
                        if isinstance(dates, str):
                            dates = [d.strip() for d in dates.split(',') if d.strip()]
                        # Rapatriement actif aujourd'hui ?
                        if today_iso not in dates:
                            continue
                        trajet = r.get('trajet', [])
                        if isinstance(trajet, str):
                            trajet = [v.strip() for v in trajet.split('>') if v.strip()]
                        dest = trajet[-1].strip() if trajet else ''
                        if dest:
                            counts[dest] = counts.get(dest, 0) + 1
                body = json.dumps(counts, ensure_ascii=False).encode('utf-8')
                self.send_response(200)
                self.send_header('Content-Type', 'application/json; charset=utf-8')
                self.send_header('Content-Length', len(body))
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(body)
            except Exception as e:
                self.send_response(500)
                self.end_headers()
                self.wfile.write(str(e).encode())

        elif self.path == '/api/messages':
            msgs = _load_messages()
            resp = json.dumps(msgs).encode()
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(resp)

        elif self.path == '/api/history':
            with _history_lock:
                body = json.dumps(_load_history(), ensure_ascii=False).encode('utf-8')
            self.send_response(200)
            self.send_header('Content-Type', 'application/json; charset=utf-8')
            self.send_header('Content-Length', len(body))
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(body)

        elif self.path == '/api/stats':
            import datetime as _dt
            try:
                with _history_lock:
                    h = _load_history()
                animals_log  = h.get('animals_log', [])
                outcomes_log = h.get('outcomes_log', [])
                year_now  = _dt.datetime.now().strftime('%Y')
                month_now = _dt.datetime.now().strftime('%Y-%m')
                # Total cette année
                total_year  = sum(1 for a in animals_log if a.get('date','').startswith(year_now))
                total_month = sum(1 for a in animals_log if a.get('date','').startswith(month_now))
                # Sorties par type
                outcomes_by_type = {}
                for o in outcomes_log:
                    if o.get('date','').startswith(year_now):
                        t = o.get('type','deces')
                        outcomes_by_type[t] = outcomes_by_type.get(t,0) + 1
                # Top espèces (animaux)
                species_count = {}
                for a in animals_log:
                    if a.get('date','').startswith(year_now):
                        s = a.get('species','') or 'Non précisée'
                        species_count[s] = species_count.get(s,0) + 1
                top_species = dict(sorted(species_count.items(), key=lambda x: -x[1])[:10])
                # Animaux par semaine (last 16 weeks)
                weekly = {}
                for a in animals_log:
                    w = a.get('week','')
                    if w:
                        weekly[w] = weekly.get(w,0) + 1
                # Générer les 16 dernières semaines dans l'ordre
                weeks_ordered = []
                for i in range(15, -1, -1):
                    d = _dt.datetime.now() - _dt.timedelta(weeks=i)
                    weeks_ordered.append(d.strftime('%Y-W%W'))
                weekly_ordered = {w: weekly.get(w, 0) for w in weeks_ordered}
                # Sorties récentes (last 100)
                recent_outcomes = sorted(outcomes_log, key=lambda x: x.get('date',''), reverse=True)[:100]
                result = {
                    'total_year': total_year,
                    'total_month': total_month,
                    'outcomes_by_type': outcomes_by_type,
                    'top_species': top_species,
                    'weekly_animals': weekly_ordered,
                    'recent_outcomes': recent_outcomes,
                }
                body = json.dumps(result, ensure_ascii=False).encode('utf-8')
                self.send_response(200)
                self.send_header('Content-Type', 'application/json; charset=utf-8')
                self.send_header('Content-Length', len(body))
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(body)
            except Exception as e:
                self.send_response(500)
                self.end_headers()
                self.wfile.write(str(e).encode())

        elif self.path == '/api/clinic-stats':
            with _clinic_stats_lock:
                body = json.dumps(_clinic_stats, ensure_ascii=False).encode('utf-8')
            self.send_response(200)
            self.send_header('Content-Type', 'application/json; charset=utf-8')
            self.send_header('Content-Length', len(body))
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(body)

        elif self.path == '/api/benevoles':
            try:
                data = get_benevoles()
                body = json.dumps(data, ensure_ascii=False).encode('utf-8')
                self.send_response(200)
                self.send_header('Content-Type', 'application/json; charset=utf-8')
                self.send_header('Content-Length', len(body))
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(body)
            except Exception as e:
                self.send_response(500)
                self.send_header('Content-Type', 'text/plain')
                self.end_headers()
                self.wfile.write(str(e).encode())

        elif self.path == '/api/benevoles/add':
            pass  # handled in do_POST

        elif self.path == '/api/vetos':
            try:
                vetos = get_vetos()
                body  = json.dumps(vetos, ensure_ascii=False).encode('utf-8')
                self.send_response(200)
                self.send_header('Content-Type', 'application/json; charset=utf-8')
                self.send_header('Content-Length', len(body))
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(body)
            except Exception as e:
                self.send_response(502)
                self.send_header('Content-Type', 'text/plain')
                self.end_headers()
                self.wfile.write(str(e).encode())

        elif self.path == '/api/config':
            body = json.dumps({
                'sheet_id': get_sheet_id()
            }).encode('utf-8')
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Content-Length', len(body))
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(body)

        elif self.path.startswith('/api/benevoles/export-xls'):
            try:
                if not _HAS_OPENPYXL:
                    raise RuntimeError('openpyxl non installé')
                # Paramètre ?year=YYYY optionnel
                year = None
                if '?year=' in self.path:
                    year = int(self.path.split('?year=')[1].split('&')[0])
                body = export_benevoles_xls(year=year)
                fname = f'benevoles_{year}.xlsx' if year else 'benevoles_export.xlsx'
                self.send_response(200)
                self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                self.send_header('Content-Disposition', f'attachment; filename="{fname}"')
                self.send_header('Content-Length', len(body))
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(body)
            except Exception as e:
                self.send_response(500)
                self.send_header('Content-Type', 'text/plain')
                self.end_headers()
                self.wfile.write(str(e).encode())

        else:
            # Pas de cache pour les fichiers HTML/JS
            if self.path.endswith('.html') or self.path == '/':
                self.send_response(200)
                self.send_header('Cache-Control', 'no-cache, no-store, must-revalidate')
                self.send_header('Pragma', 'no-cache')
                self.send_header('Expires', '0')
                import mimetypes
                ctype, _ = mimetypes.guess_type(self.path.lstrip('/') or 'index.html')
                self.send_header('Content-Type', ctype or 'text/html; charset=utf-8')
                self.end_headers()
                fname = self.path.lstrip('/') or 'heatmap.html'
                try:
                    with open(fname, 'rb') as f:
                        self.wfile.write(f.read())
                except FileNotFoundError:
                    pass
            else:
                super().do_GET()

    def do_POST(self):
        if self.path == '/api/geocache':
            try:
                length = int(self.headers.get('Content-Length', 0))
                body   = json.loads(self.rfile.read(length))
                ville  = str(body['ville']).strip()
                lat    = float(body['lat'])
                lon    = float(body['lon'])
                ok, msg = add_to_geo_preloaded(ville, lat, lon)
                resp = json.dumps({'ok': ok, 'msg': msg}).encode()
                self.send_response(200 if ok else 400)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(resp)
            except Exception as e:
                self.send_response(500)
                self.send_header('Content-Type', 'text/plain')
                self.end_headers()
                self.wfile.write(str(e).encode())

        elif self.path == '/api/clinic-stats':
            try:
                length = int(self.headers.get('Content-Length', 0))
                body   = json.loads(self.rfile.read(length))
                code   = str(body.get('code', '')).strip().upper()
                if code:
                    with _clinic_stats_lock:
                        _clinic_stats[code] = {
                            'animaux':      int(body.get('animaux', 0)),
                            'deces':        int(body.get('deces', 0)),
                            'depots':       int(body.get('depots', 0)),
                            'enRoute':      int(body.get('enRoute', 0)),
                            'lastSeen':     int(body.get('lastSeen', 0)),
                            'sheetArrived': list(body.get('sheetArrived', [])),
                            'animalsList':  list(body.get('animalsList', [])),
                            'outcomesList': list(body.get('outcomesList', [])),
                            'depotsList':   list(body.get('depotsList', [])),
                            'ts':           time.time(),
                        }
                        _save_clinic_stats(_clinic_stats)
                        _log_new_entries(code, body)
                resp = json.dumps({'ok': True}).encode()
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(resp)
            except Exception as e:
                self.send_response(500)
                self.send_header('Content-Type', 'text/plain')
                self.end_headers()
                self.wfile.write(str(e).encode())

        elif self.path == '/api/clinic-stats/delete-item':
            try:
                length = int(self.headers.get('Content-Length', 0))
                body   = json.loads(self.rfile.read(length))
                code   = str(body.get('code', '')).strip().upper()
                liste  = str(body.get('liste', ''))   # animalsList / outcomesList / depotsList
                item_id = body.get('id')
                if code and liste in ('animalsList','outcomesList','depotsList') and item_id is not None:
                    with _clinic_stats_lock:
                        if code in _clinic_stats and liste in _clinic_stats[code]:
                            _clinic_stats[code][liste] = [
                                x for x in _clinic_stats[code][liste]
                                if str(x.get('id')) != str(item_id)
                            ]
                            # Recalculer le compteur
                            count_map = {'animalsList':'animaux','outcomesList':'deces','depotsList':'depots'}
                            key = count_map.get(liste)
                            if key:
                                _clinic_stats[code][key] = len(_clinic_stats[code][liste])
                            _save_clinic_stats(_clinic_stats)
                resp = json.dumps({'ok': True}).encode()
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(resp)
            except Exception as e:
                self.send_response(500)
                self.send_header('Content-Type', 'text/plain')
                self.end_headers()
                self.wfile.write(str(e).encode())

        elif self.path == '/api/benevoles/add':
            try:
                length = int(self.headers.get('Content-Length', 0))
                body   = json.loads(self.rfile.read(length))
                add_benevole(body)
                resp = json.dumps({'ok': True}).encode()
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(resp)
            except Exception as e:
                self.send_response(500)
                self.send_header('Content-Type', 'text/plain')
                self.end_headers()
                self.wfile.write(str(e).encode())

        elif self.path == '/api/benevoles/update':
            try:
                length = int(self.headers.get('Content-Length', 0))
                body   = json.loads(self.rfile.read(length))
                orig_nom    = str(body.pop('_orig_nom', '')).strip()
                orig_prenom = str(body.pop('_orig_prenom', '')).strip()
                update_benevole(orig_nom, orig_prenom, body)
                resp = json.dumps({'ok': True}).encode()
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(resp)
            except ValueError as e:
                resp = str(e).encode()
                self.send_response(404)
                self.send_header('Content-Type', 'text/plain')
                self.end_headers()
                self.wfile.write(resp)
            except Exception as e:
                self.send_response(500)
                self.send_header('Content-Type', 'text/plain')
                self.end_headers()
                self.wfile.write(str(e).encode())

        elif self.path == '/api/config':
            try:
                length   = int(self.headers.get('Content-Length', 0))
                body     = json.loads(self.rfile.read(length))
                new_id   = str(body.get('sheet_id', '')).strip()
                if not new_id:
                    raise ValueError('sheet_id manquant')
                set_sheet_id(new_id)
                resp = json.dumps({'ok': True, 'sheet_id': new_id}).encode()
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(resp)
            except Exception as e:
                self.send_response(400)
                self.send_header('Content-Type', 'text/plain')
                self.end_headers()
                self.wfile.write(str(e).encode())

        elif self.path == '/api/benevoles/reimport-xls':
            try:
                xls_data = get_benevoles_from_xls()
                with _benv_lock:
                    existing = load_benevoles_json() or []
                    # Index existing by nom+prenom
                    idx = {(b.get('nom','').strip(), b.get('prenom','').strip()): i for i, b in enumerate(existing)}
                    added = 0
                    updated = 0
                    for xb in xls_data:
                        key = (xb.get('nom','').strip(), xb.get('prenom','').strip())
                        if key in idx:
                            # Update XLS fields only, keep extra fields
                            i = idx[key]
                            for k in ['roles','telephone','email','commune','rayon','naissance','immat','puissance','moteur','competences','commentaire','info_temp','derniers_rapat','date_creation','modif']:
                                existing[i][k] = xb.get(k, existing[i].get(k,''))
                            updated += 1
                        else:
                            xb['_id'] = str(_uuid.uuid4())
                            for k in EXTRA_FIELDS:
                                xb.setdefault(k, '')
                            existing.append(xb)
                            added += 1
                    save_benevoles_json(existing)
                resp = json.dumps({'ok': True, 'added': added, 'updated': updated}).encode()
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(resp)
            except Exception as e:
                self.send_response(500)
                self.send_header('Content-Type', 'text/plain')
                self.end_headers()
                self.wfile.write(str(e).encode())

        elif self.path == '/api/benevoles/clear-rapatriements':
            try:
                length = int(self.headers.get('Content-Length', 0))
                body   = json.loads(self.rfile.read(length))
                year   = int(body['year'])
                result = clear_rapatriements_year(year)
                resp   = json.dumps({'ok': True, **result}).encode()
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(resp)
            except Exception as e:
                self.send_response(500)
                self.send_header('Content-Type', 'text/plain')
                self.end_headers()
                self.wfile.write(str(e).encode())

        elif self.path == '/api/rapatriements/add':
            try:
                length = int(self.headers.get('Content-Length', 0))
                body   = json.loads(self.rfile.read(length))
                benv_id     = str(body.get('benevole_id', '')).strip()
                benv_nom    = str(body.get('benevole_nom', '')).strip()
                benv_prenom = str(body.get('benevole_prenom', '')).strip()
                dates       = body.get('dates', [])
                if len(dates) == 1:
                    date_label = dates[0]
                elif len(dates) > 1:
                    date_label = f"{dates[0]} - {dates[-1]}"
                else:
                    date_label = ''
                # Formater pour l'affichage (dd/mm/yyyy) mais garder ISO dans 'dates'
                def to_display(d):
                    parts = d.split('-')
                    if len(parts) == 3 and len(parts[0]) == 4:
                        return f"{parts[2]}/{parts[1]}/{parts[0]}"
                    return d
                if len(dates) == 1:
                    date_label = to_display(dates[0])
                elif len(dates) > 1:
                    date_label = f"{to_display(dates[0])} - {to_display(dates[-1])}"
                else:
                    date_label = ''
                rapat_entry = {
                    'date':          date_label,   # format dd/mm/yyyy pour affichage
                    'dates':         dates,         # format ISO yyyy-mm-dd pour comparaisons
                    'heure_depart':  str(body.get('heure_depart', '')),
                    'heure_arrivee': str(body.get('heure_arrivee', '')),
                    'trajet':        body.get('trajet', []),
                    'distance_km':   str(body.get('distance_km', '')),
                    'notes':         str(body.get('notes', '')),
                }
                with _benv_lock:
                    volunteers = load_benevoles_json() or []
                    found = False
                    for b in volunteers:
                        if benv_id and b.get('_id', '') == benv_id:
                            found = True
                        elif not benv_id and b.get('nom','').strip() == benv_nom and b.get('prenom','').strip() == benv_prenom:
                            found = True
                        if found:
                            raw = b.get('derniers_rapat', '')
                            if isinstance(raw, str):
                                # Convertir string en liste
                                if raw.strip():
                                    existing_list = []
                                    for ligne in raw.split('\n'):
                                        ligne = ligne.strip()
                                        if not ligne:
                                            continue
                                        if ' : ' in ligne:
                                            d, t = ligne.split(' : ', 1)
                                        else:
                                            d, t = '', ligne
                                        existing_list.append({'date': d.strip(), 'trajet': t.strip()})
                                else:
                                    existing_list = []
                                b['derniers_rapat'] = existing_list
                            elif not isinstance(b['derniers_rapat'], list):
                                b['derniers_rapat'] = []
                            b['derniers_rapat'].append(rapat_entry)
                            break
                    if not found:
                        raise ValueError(f'Bénévole non trouvé : {benv_prenom} {benv_nom}')
                    save_benevoles_json(volunteers)
                print(f'[rapatriements] ✓ Ajouté pour {benv_prenom} {benv_nom} : {date_label}')
                resp = json.dumps({'ok': True}).encode()
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(resp)
            except ValueError as e:
                self.send_response(404)
                self.send_header('Content-Type', 'text/plain')
                self.end_headers()
                self.wfile.write(str(e).encode())
            except Exception as e:
                self.send_response(500)
                self.send_header('Content-Type', 'text/plain')
                self.end_headers()
                self.wfile.write(str(e).encode())

        elif self.path == '/api/benevoles/delete':
            try:
                length  = int(self.headers.get('Content-Length', 0))
                body    = json.loads(self.rfile.read(length))
                benv_id = str(body.get('benevole_id', ''))
                with _benv_lock:
                    volunteers = load_benevoles_json() or []
                    before = len(volunteers)
                    volunteers = [b for b in volunteers if b.get('_id') != benv_id]
                    deleted = before - len(volunteers)
                    if deleted:
                        save_benevoles_json(volunteers)
                resp = json.dumps({'ok': deleted > 0, 'deleted': deleted}).encode()
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(resp)
            except Exception as e:
                self.send_response(500)
                self.send_header('Content-Type', 'text/plain')
                self.end_headers()
                self.wfile.write(str(e).encode())

        elif self.path == '/api/rapatriements/delete':
            try:
                length = int(self.headers.get('Content-Length', 0))
                body   = json.loads(self.rfile.read(length))
                benv_id = str(body.get('benevole_id', ''))
                idx     = int(body.get('index', -1))
                with _benv_lock:
                    volunteers = load_benevoles_json() or []
                    found = False
                    for b in volunteers:
                        if b.get('_id') == benv_id:
                            raw = b.get('derniers_rapat', '')
                            # Convertir texte en liste si besoin
                            if isinstance(raw, str):
                                items = []
                                for ligne in raw.split('\n'):
                                    ligne = ligne.strip()
                                    if not ligne: continue
                                    if ' : ' in ligne:
                                        d, t = ligne.split(' : ', 1)
                                    else:
                                        d, t = '', ligne
                                    items.append({'date': d.strip(), 'trajet': t.strip()})
                                b['derniers_rapat'] = items
                            if 0 <= idx < len(b['derniers_rapat']):
                                b['derniers_rapat'].pop(idx)
                                found = True
                            break
                    if found:
                        save_benevoles_json(volunteers)
                resp = json.dumps({'ok': found}).encode()
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(resp)
            except Exception as e:
                self.send_response(500)
                self.send_header('Content-Type', 'text/plain')
                self.end_headers()
                self.wfile.write(str(e).encode())

        elif self.path == '/api/rapatriements/confirm':
            try:
                from datetime import datetime as _dt
                length  = int(self.headers.get('Content-Length', 0))
                body    = json.loads(self.rfile.read(length))
                benv_id = str(body.get('benevole_id', ''))
                idx     = int(body.get('index', -1))
                done    = bool(body.get('done', False))
                with _benv_lock:
                    volunteers = load_benevoles_json() or []
                    found = False
                    for b in volunteers:
                        if b.get('_id') == benv_id:
                            raw = b.get('derniers_rapat', [])
                            if isinstance(raw, list) and 0 <= idx < len(raw):
                                raw[idx]['confirme']    = done
                                raw[idx]['confirme_at'] = _dt.now().strftime('%d/%m/%y à %H:%M')
                                found = True
                            break
                    if found:
                        save_benevoles_json(volunteers)
                resp = json.dumps({'ok': found}).encode()
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(resp)
            except Exception as e:
                self.send_response(500)
                self.send_header('Content-Type', 'text/plain')
                self.end_headers()
                self.wfile.write(str(e).encode())

        elif self.path == '/api/messages':
            try:
                length = int(self.headers.get('Content-Length', 0))
                body = json.loads(self.rfile.read(length))
                msgs = _load_messages()
                msg = {
                    'id': _secrets.token_hex(8),
                    'clinic_code': _sanitize(body.get('clinic_code', ''), 20),
                    'clinic_name': _sanitize(body.get('clinic_name', ''), 100),
                    'text':        _sanitize(body.get('text', ''), 2000).strip(),
                    'timestamp':   _sanitize(body.get('timestamp', ''), 50),
                    'read': False
                }
                if not msg['text']:
                    self.send_response(400)
                    self.end_headers()
                    return
                msgs.insert(0, msg)
                _save_messages(msgs)
                resp = json.dumps({'ok': True}).encode()
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(resp)
            except Exception as e:
                self.send_response(500)
                self.end_headers()
                self.wfile.write(str(e).encode())

        elif self.path == '/api/messages/read':
            try:
                length = int(self.headers.get('Content-Length', 0))
                body = json.loads(self.rfile.read(length))
                msg_id = str(body.get('id', ''))
                msgs = _load_messages()
                for m in msgs:
                    if m['id'] == msg_id:
                        m['read'] = True
                        break
                _save_messages(msgs)
                resp = json.dumps({'ok': True}).encode()
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(resp)
            except Exception as e:
                self.send_response(500)
                self.end_headers()
                self.wfile.write(str(e).encode())

        elif self.path == '/api/auth':
            try:
                ip = self.client_address[0]
                allowed, retry_after = _check_rate_limit(ip)
                if not allowed:
                    resp = json.dumps({'ok': False, 'error': 'too_many_attempts', 'retry_after': retry_after}).encode()
                    self.send_response(429)
                    self.send_header('Content-Type', 'application/json')
                    self.send_header('Retry-After', str(retry_after))
                    self.send_header('Access-Control-Allow-Origin', _cors_origin())
                    self.end_headers()
                    self.wfile.write(resp)
                    return
                length = int(self.headers.get('Content-Length', 0))
                body   = json.loads(self.rfile.read(length))
                if _check_portal_password(str(body.get('password', ''))):
                    token = _secrets.token_hex(32)
                    _add_token(token)
                    _record_login_success(ip)
                    resp = json.dumps({'ok': True, 'token': token}).encode()
                    self.send_response(200)
                else:
                    _record_login_failure(ip)
                    with _login_lock:
                        remaining = MAX_LOGIN_ATTEMPTS - _login_attempts.get(ip, {}).get('count', 0)
                    resp = json.dumps({'ok': False, 'remaining': max(0, remaining)}).encode()
                    self.send_response(401)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', _cors_origin())
                self.end_headers()
                self.wfile.write(resp)
            except Exception as e:
                self.send_response(500)
                self.end_headers()
                self.wfile.write(str(e).encode())

        elif self.path == '/api/auth/verify':
            try:
                length = int(self.headers.get('Content-Length', 0))
                body   = json.loads(self.rfile.read(length))
                token  = str(body.get('token', ''))
                ok     = _check_token(token)
                resp   = json.dumps({'ok': ok}).encode()
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', _cors_origin())
                self.end_headers()
                self.wfile.write(resp)
            except Exception as e:
                self.send_response(500)
                self.end_headers()
                self.wfile.write(str(e).encode())

        else:
            self.send_response(404)
            self.end_headers()

    def log_message(self, fmt, *args):
        if '/api/' in args[0]:
            print(f'[{args[1]}] {args[0]}')


if __name__ == '__main__':
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    # Pré-chargement au démarrage
    threading.Thread(target=get_data, daemon=True).start()

    import socket
    hostname = socket.gethostname()
    try:
        local_ip = socket.gethostbyname(hostname)
    except Exception:
        local_ip = '127.0.0.1'

    sheet_id = get_sheet_id()
    print(f'\n🗺️  Heatmap Animaux Sauvages')
    print(f'   http://localhost:{PORT}')
    print(f'   http://{local_ip}:{PORT}  ← partager cette adresse')
    print(f'   Sheet ID : {sheet_id}\n')

    with http.server.ThreadingHTTPServer(('', PORT), Handler) as srv:
        try:
            srv.serve_forever()
        except KeyboardInterrupt:
            print('\nServeur arrêté.')
